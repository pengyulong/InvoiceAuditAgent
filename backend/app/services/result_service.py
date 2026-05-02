"""结果服务 - 审计结果查询和报告导出"""
import os
import json
import logging
import tempfile
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta

from app.services.audit_service import audit_service

logger = logging.getLogger(__name__)


class ResultService:
    """审计结果查询和报告导出服务"""

    async def get_result_details(self, audit_id: str) -> Optional[Dict[str, Any]]:
        """获取审计结果详情"""
        return await audit_service.get_audit_results(audit_id)

    async def get_result_summary(self, audit_id: str) -> Optional[Dict[str, Any]]:
        """获取审计结果摘要"""
        result = await self.get_result_details(audit_id)
        if not result:
            return None
        summary = result.get("summary", {})
        return {
            "audit_id": audit_id,
            "summary": summary,
            "issue_count": len(result.get("issues", [])),
            "invoice_count": len(result.get("invoices", [])),
            "status": summary.get("status", "unknown"),
        }

    async def get_audit_issues(self, audit_id: str, severity: Optional[str] = None,
                                type: Optional[str] = None) -> List[Dict[str, Any]]:
        result = await self.get_result_details(audit_id)
        if not result:
            return []
        issues = result.get("issues", [])
        if severity:
            issues = [i for i in issues if i.get("severity") == severity]
        if type:
            issues = [i for i in issues if i.get("type") == type]
        return issues

    async def get_comparisons(self, audit_id: str) -> List[Dict[str, Any]]:
        result = await self.get_result_details(audit_id)
        return result.get("comparisons", []) if result else []

    async def get_contract_info(self, audit_id: str) -> Optional[Dict[str, Any]]:
        result = await self.get_result_details(audit_id)
        return result.get("contract_info") if result else None

    async def get_invoice_info(self, audit_id: str, status: Optional[str] = None) -> List[Dict[str, Any]]:
        result = await self.get_result_details(audit_id)
        if not result:
            return []
        invoices = result.get("invoices", [])
        if status:
            invoices = [inv for inv in invoices if inv.get("status") == status]
        return invoices

    async def generate_pdf_report(self, audit_id: str, include_details: bool = True) -> str:
        """生成PDF审计报告"""
        result = await self.get_result_details(audit_id)
        if not result:
            raise ValueError("审计结果不存在")

        os.makedirs("uploads/reports", exist_ok=True)
        pdf_path = f"uploads/reports/audit_report_{audit_id}.pdf"

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # 尝试注册中文字体
            font_registered = False
            for font_path in [
                "/System/Library/Fonts/PingFang.ttc",
                "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
                "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            ]:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont("Chinese", font_path))
                        font_registered = True
                        break
                    except Exception:
                        continue

            doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                                    leftMargin=20*mm, rightMargin=20*mm,
                                    topMargin=20*mm, bottomMargin=20*mm)
            styles = getSampleStyleSheet()

            if font_registered:
                body_style = ParagraphStyle("ChineseBody", parent=styles["Normal"],
                                            fontName="Chinese", fontSize=10, leading=16)
                title_style = ParagraphStyle("ChineseTitle", parent=styles["Title"],
                                             fontName="Chinese", fontSize=18)
                heading_style = ParagraphStyle("ChineseHeading", parent=styles["Heading2"],
                                               fontName="Chinese", fontSize=14)
            else:
                body_style = styles["Normal"]
                title_style = styles["Title"]
                heading_style = styles["Heading2"]

            story = []
            summary = result.get("summary", {})

            story.append(Paragraph("智能合同发票审计报告", title_style))
            story.append(Spacer(1, 12*mm))

            story.append(Paragraph("一、审计摘要", heading_style))
            story.append(Paragraph(
                f"审计结论: {self._cn_status(summary.get('status', 'unknown'))} | "
                f"合同金额: ¥{summary.get('contract_amount', 0):,.2f} | "
                f"发票总额: ¥{summary.get('invoice_total', 0):,.2f} | "
                f"覆盖率: {summary.get('coverage', 0):.1%} | "
                f"问题数: {summary.get('issue_count', 0)} | "
                f"置信度: {(summary.get('confidence_score', 0) or 0):.1%}",
                body_style
            ))
            story.append(Spacer(1, 8*mm))

            # 合同信息
            contract = result.get("contract_info") or {}
            if contract:
                story.append(Paragraph("二、合同信息", heading_style))
                contract_data = [
                    ["合同编号", str(contract.get("contract_number", "-"))],
                    ["买方", str(contract.get("buyer_name", "-"))],
                    ["卖方", str(contract.get("seller_name", "-"))],
                    ["合同金额", f"¥{contract.get('total_amount', 0) or 0:,.2f}"],
                    ["税率", f"{(contract.get('tax_rate', 0) or 0) * 100:.0f}%"],
                    ["置信度", f"{(contract.get('confidence_score', 0) or 0):.1%}"],
                ]
                t = Table(contract_data, colWidths=[80*mm, 80*mm])
                t.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.95)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTNAME", (0, 0), (-1, -1),
                     "Chinese" if font_registered else "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(t)
                story.append(Spacer(1, 8*mm))

            # 发票列表
            invoices = result.get("invoices", [])
            if invoices:
                story.append(Paragraph("三、发票列表", heading_style))
                inv_header = ["发票号码", "金额", "状态", "置信度"]
                inv_data = [inv_header]
                for inv in invoices[:20]:
                    inv_data.append([
                        str(inv.get("invoice_number", "-")),
                        f"¥{inv.get('total_amount', 0) or 0:,.2f}",
                        inv.get("status", "normal"),
                        f"{(inv.get('confidence_score', 0) or 0):.1%}",
                    ])
                t = Table(inv_data, colWidths=[45*mm, 40*mm, 35*mm, 40*mm])
                t.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.6)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTNAME", (0, 0), (-1, -1),
                     "Chinese" if font_registered else "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]))
                story.append(t)
                story.append(Spacer(1, 8*mm))

            # 问题列表
            issues = result.get("issues", [])
            if issues:
                story.append(Paragraph("四、发现的问题", heading_style))
                for i, issue in enumerate(issues, 1):
                    sev_label = {"high": "严重", "medium": "中等", "low": "轻微"}
                    sev = sev_label.get(issue.get("severity", ""), "未知")
                    story.append(Paragraph(
                        f"<b>{i}. [{sev}] {issue.get('title', '')}</b>", body_style
                    ))
                    story.append(Paragraph(
                        f"描述: {issue.get('description', '')}<br/>"
                        f"建议: {issue.get('recommendation', '')}",
                        body_style
                    ))
                    story.append(Spacer(1, 3*mm))

            story.append(Spacer(1, 10*mm))
            story.append(Paragraph(
                f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                body_style
            ))

            doc.build(story)
            logger.info(f"PDF报告生成成功: {pdf_path}")

        except ImportError:
            logger.warning("reportlab未安装，生成简化PDF")
            with open(pdf_path, "w", encoding="utf-8") as f:
                f.write(self._text_report(result))

        return pdf_path

    async def generate_excel_report(self, audit_id: str, include_raw_data: bool = False) -> str:
        """生成Excel审计报告"""
        result = await self.get_result_details(audit_id)
        if not result:
            raise ValueError("审计结果不存在")

        os.makedirs("uploads/reports", exist_ok=True)
        excel_path = f"uploads/reports/audit_report_{audit_id}.xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            header_fill = PatternFill(start_color="3355CC", end_color="3355CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            # Sheet 1: 摘要
            ws = wb.active
            ws.title = "审计摘要"
            summary = result.get("summary", {})
            ws["A1"] = "智能合同发票审计报告"
            ws["A1"].font = Font(size=16, bold=True)
            ws.merge_cells("A1:D1")

            summary_data = [
                ["审计结论", self._cn_status(summary.get("status", "unknown"))],
                ["合同金额", f"¥{summary.get('contract_amount', 0):,.2f}"],
                ["发票总额", f"¥{summary.get('invoice_total', 0):,.2f}"],
                ["覆盖率", f"{summary.get('coverage', 0):.1%}"],
                ["问题数", summary.get("issue_count", 0)],
                ["处理时间", f"{summary.get('processing_time', 0):.1f}秒"],
                ["置信度", f"{(summary.get('confidence_score', 0) or 0):.1%}"],
            ]
            for i, (label, value) in enumerate(summary_data, 3):
                ws.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=i, column=2, value=str(value))

            # Sheet 2: 合同信息
            ws2 = wb.create_sheet("合同信息")
            contract = result.get("contract_info") or {}
            contract_fields = [
                ("合同编号", "contract_number"), ("买方", "buyer_name"),
                ("卖方", "seller_name"), ("合同金额", "total_amount"),
                ("税率", "tax_rate"), ("合同日期", "contract_date"),
                ("置信度", "confidence_score"),
            ]
            for i, (label, key) in enumerate(contract_fields, 1):
                ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
                val = contract.get(key, "-")
                if key == "tax_rate":
                    try:
                        if isinstance(val, str):
                            val = float(val.replace("%", ""))
                        val = f"{(val or 0) * 100:.0f}%"
                    except (ValueError, TypeError):
                        val = str(val)
                elif key == "total_amount":
                    val = f"¥{(val or 0):,.2f}"
                ws2.cell(row=i, column=2, value=str(val))

            # Sheet 3: 发票列表
            ws3 = wb.create_sheet("发票列表")
            inv_headers = ["发票号码", "金额", "税额", "状态", "置信度", "是否重复"]
            for col, h in enumerate(inv_headers, 1):
                cell = ws3.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for i, inv in enumerate(result.get("invoices", []), 2):
                row_data = [
                    inv.get("invoice_number", "-"),
                    inv.get("total_amount", 0) or 0,
                    inv.get("tax_amount", 0) or 0,
                    inv.get("status", "normal"),
                    (inv.get("confidence_score", 0) or 0),
                    "是" if inv.get("is_duplicate") else "否",
                ]
                for col, val in enumerate(row_data, 1):
                    cell = ws3.cell(row=i, column=col, value=val)
                    cell.border = thin_border

            # Sheet 4: 问题清单
            ws4 = wb.create_sheet("问题清单")
            issue_headers = ["严重程度", "类型", "标题", "描述", "建议"]
            for col, h in enumerate(issue_headers, 1):
                cell = ws4.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            for i, issue in enumerate(result.get("issues", []), 2):
                row_data = [
                    issue.get("severity", "-"),
                    issue.get("type", "-"),
                    issue.get("title", "-"),
                    issue.get("description", "-"),
                    issue.get("recommendation", "-"),
                ]
                for col, val in enumerate(row_data, 1):
                    cell = ws4.cell(row=i, column=col, value=val)
                    if issue.get("severity") == "high":
                        cell.fill = warn_fill

            wb.save(excel_path)
            logger.info(f"Excel报告生成成功: {excel_path}")

        except ImportError:
            logger.warning("openpyxl未安装，生成简化Excel")
            with open(excel_path, "w", encoding="utf-8") as f:
                f.write(self._text_report(result))

        return excel_path

    async def generate_invoice_excel(self, task_id: str) -> str:
        """生成发票识别Excel（单sheet，仅发票字段）"""
        result = await self.get_result_details(task_id)
        if not result:
            raise ValueError("审计结果不存在")

        os.makedirs("uploads/reports", exist_ok=True)
        excel_path = f"uploads/reports/invoice_report_{task_id}.xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "发票信息"

            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            header_fill = PatternFill(start_color="3355CC", end_color="3355CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            headers = [
                "发票号码", "开票日期", "销售方", "购买方",
                "价税合计", "税额", "不含税金额", "税率",
                "商品清单", "置信度", "来源文件"
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = header_align

            invoices = result.get("invoices", []) or result.get("invoice_list", [])
            for i, inv in enumerate(invoices, 2):
                product_list = inv.get("product_list") or []
                if isinstance(product_list, list) and product_list:
                    items = []
                    for p in product_list:
                        if isinstance(p, dict):
                            name = p.get("name", "")
                            qty = p.get("quantity", "")
                            items.append(f"{name} x{qty}" if qty else name)
                    product_text = "; ".join(items)
                else:
                    product_text = "-"

                tax_rate = inv.get("tax_rate")
                if tax_rate is not None:
                    try:
                        if isinstance(tax_rate, str):
                            tax_rate = tax_rate.replace("%", "")
                        tax_rate_val = float(tax_rate)
                        tax_rate_text = f"{tax_rate_val * 100:.0f}%" if tax_rate_val < 1 else f"{tax_rate_val:.0f}%"
                    except (ValueError, TypeError):
                        tax_rate_text = str(tax_rate)
                else:
                    tax_rate_text = "-"

                source = inv.get("source_file", inv.get("_source", "-"))

                row_data = [
                    inv.get("invoice_number") or "-",
                    inv.get("invoice_date") or "-",
                    inv.get("seller_name") or "-",
                    inv.get("buyer_name") or "-",
                    inv.get("total_amount") or 0,
                    inv.get("tax_amount") or 0,
                    inv.get("amount_without_tax") or 0,
                    tax_rate_text,
                    product_text,
                    (inv.get("confidence_score") or 0),
                    source,
                ]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.border = thin_border
                    if col == 10:
                        cell.number_format = '0.0%'

            # 列宽自适应
            col_widths = [18, 14, 22, 22, 14, 14, 14, 10, 40, 10, 30]
            for col, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

            wb.save(excel_path)
            logger.info(f"发票Excel生成成功: {excel_path}")

        except ImportError:
            logger.warning("openpyxl未安装，生成简化发票Excel")
            with open(excel_path, "w", encoding="utf-8") as f:
                f.write(self._text_report(result))

        return excel_path

    async def export_json_data(self, audit_id: str, include_images: bool = False) -> Dict[str, Any]:
        result = await self.get_result_details(audit_id)
        return {
            "export_time": datetime.now().isoformat(),
            "audit_id": audit_id,
            "include_images": include_images,
            "data": result,
            "metadata": {"export_version": "1.0.0", "format": "json", "schema_version": "1.0"},
        }

    async def get_image_file(self, audit_id: str, image_type: str, image_id: str) -> str:
        return f"uploads/images/{audit_id}/{image_type}/{image_id}.jpg"

    async def generate_share_link(self, audit_id: str, expire_hours: int = 24,
                                   password: Optional[str] = None) -> Dict[str, Any]:
        share_id = f"share_{audit_id}_{int(datetime.now().timestamp())}"
        return {
            "share_id": share_id,
            "audit_id": audit_id,
            "expires_at": (datetime.now() + timedelta(hours=expire_hours)).isoformat(),
            "has_password": password is not None,
            "share_url": f"/results/shared/{share_id}",
            "created_at": datetime.now().isoformat(),
        }

    async def get_shared_result(self, share_id: str, password: Optional[str] = None) -> Optional[Dict[str, Any]]:
        if not share_id.startswith("share_"):
            return None
        parts = share_id.split("_")
        if len(parts) < 2:
            return None
        audit_id = parts[1]
        return await self.get_result_details(audit_id)

    @staticmethod
    def _cn_status(status: str) -> str:
        status_map = {"pass": "通过", "fail": "不通过", "warning": "需人工复核", "unknown": "未知"}
        return status_map.get(status, status)

    def _text_report(self, result: Dict) -> str:
        """纯文本报告（富文本库不可用时）"""
        lines = ["智能合同发票审计报告", "=" * 30]
        summary = result.get("summary", {})
        lines.append(f"审计结论: {self._cn_status(summary.get('status', 'unknown'))}")
        lines.append(f"合同金额: ¥{summary.get('contract_amount', 0):,.2f}")
        lines.append(f"发票总额: ¥{summary.get('invoice_total', 0):,.2f}")
        lines.append(f"覆盖率: {summary.get('coverage', 0):.1%}")
        lines.append(f"问题数: {summary.get('issue_count', 0)}")
        lines.append(f"报告生成: {datetime.now().isoformat()}")
        return "\n".join(lines)


result_service = ResultService()
